import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from config import settings

def generate_report(registros: list, curso: str, fecha: str, docente: str = "") -> str:
    Path(settings.REPORTS_DIR).mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"asistencia_{curso}_{fecha}_{timestamp}.xlsx"
    filepath = Path(settings.REPORTS_DIR) / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    # ── Colores ──────────────────────────────────────────
    COLOR_HEADER_BG   = "1F3864"   # Azul oscuro
    COLOR_HEADER_FONT = "FFFFFF"   # Blanco
    COLOR_TITLE_BG    = "2E75B6"   # Azul medio
    COLOR_ROW_ALT     = "DCE6F1"   # Azul muy claro (filas alternas)
    COLOR_BORDER      = "B8CCE4"

    thin = Side(style="thin", color=COLOR_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Fila 1: Título principal ──────────────────────────
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "REPORTE DE ASISTENCIA"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color=COLOR_HEADER_FONT)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Filas 2-4: Metadata ───────────────────────────────
    meta = [
        ("Curso:",   curso),
        ("Fecha:",   fecha),
        ("Docente:", docente),
    ]
    for i, (label, value) in enumerate(meta, start=2):
        ws.merge_cells(f"A{i}:B{i}")
        ws.merge_cells(f"C{i}:E{i}")
        label_cell = ws[f"A{i}"]
        value_cell = ws[f"C{i}"]
        label_cell.value = label
        value_cell.value = value
        label_cell.font = Font(name="Calibri", bold=True, size=11, color=COLOR_HEADER_BG)
        value_cell.font = Font(name="Calibri", size=11)
        label_cell.alignment = Alignment(horizontal="right", vertical="center")
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[i].height = 18

    # ── Fila 5: Encabezados de tabla ──────────────────────
    headers = ["#", "Código", "Nombre", "Fecha", "Hora"]
    header_row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(name="Calibri", bold=True, size=11, color=COLOR_HEADER_FONT)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[header_row].height = 22

    # ── Filas de datos ────────────────────────────────────
    for idx, registro in enumerate(registros, start=1):
        row = header_row + idx
        alt_fill = PatternFill("solid", fgColor=COLOR_ROW_ALT) if idx % 2 == 0 else None
        values = [
            idx,
            registro.get("codigo", ""),
            registro.get("nombre", ""),
            registro.get("fecha", ""),
            registro.get("hora", ""),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(horizontal="center" if col != 3 else "left", vertical="center")
            cell.border = border
            if alt_fill:
                cell.fill = alt_fill
        ws.row_dimensions[row].height = 18

    # ── Fila final: Total ─────────────────────────────────
    total_row = header_row + len(registros) + 1
    ws.merge_cells(f"A{total_row}:D{total_row}")
    ws.cell(row=total_row, column=1, value="Total asistentes:").font = Font(bold=True, size=11, color=COLOR_HEADER_BG)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=total_row, column=5, value=len(registros))
    total_cell.font = Font(bold=True, size=12, color=COLOR_HEADER_BG)
    total_cell.alignment = Alignment(horizontal="center")

    # ── Ancho de columnas ─────────────────────────────────
    col_widths = [6, 20, 30, 15, 12]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(filepath)
    return str(filepath)